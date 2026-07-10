"""
Document Loader — Ingests university documents into Azure AI Search.

Supports: PDF, DOCX, DOC, XLSX, XLS, CSV, TXT, MD

Run this as a script to process documents and upload them to the vector index.
Usage: python -m app.services.document_loader --path ./data/sample/
"""

import os
import io
import csv
import time
import hashlib
import argparse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from openai import AzureOpenAI
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SearchField,
    SearchFieldDataType,
    SimpleField,
    SearchableField,
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
)
from azure.core.credentials import AzureKeyCredential

from pypdf import PdfReader
from docx import Document

from app.config import get_settings

settings = get_settings()

# All supported file extensions
SUPPORTED_EXTENSIONS = (
    ".pdf", ".docx", ".doc",
    ".xlsx", ".xls",
    ".csv",
    ".txt", ".md",
)


def chunk_text(text: str, chunk_size: int = 600, overlap: int = 100) -> list[str]:
    """Split text into overlapping chunks for better retrieval.

    Defaults reduced from 1000/200 to 600/100 to produce tighter,
    more semantically focused chunks — improves precision after
    preprocessing removed boilerplate (headers, footers, nav).
    """
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk.strip())
        start += chunk_size - overlap
    return chunks


def extract_text_from_pdf(file_path: str) -> str:
    try:
        reader = PdfReader(file_path)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as e:
        print(f"  Warning: Failed to extract text from {file_path}: {e}")
        return ""


def extract_text_from_docx(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join(para.text for para in doc.paragraphs)


def extract_text_from_excel(file_path: str) -> str:
    """Extract text from an Excel workbook (xlsx/xls), converting each sheet to CSV.

    Handles:
      - Multiple sheets (separated by '--- Sheet: <name> ---')
      - Merged cells (fills merged region with top-left cell value)
      - Formula cells (reads cached computed values via data_only=True)
      - Preserves column structure via CSV formatting
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".xls":
        return _extract_text_from_xls(file_path)
    else:
        return _extract_text_from_xlsx(file_path)


def _extract_text_from_xlsx(file_path: str) -> str:
    """Extract text from .xlsx files using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print(f"  Warning: openpyxl not installed. Cannot process {file_path}")
        return ""

    sheets_text: list[str] = []

    try:
        # data_only=True reads cached formula results instead of formula strings
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Build a dict of merged cell values so we can fill the merged regions
            merged_values: dict[tuple[int, int], str] = {}
            for merge_range in ws.merged_cells.ranges:
                # Get the value from the top-left cell of the merged region
                top_left_value = ws.cell(
                    row=merge_range.min_row, column=merge_range.min_col
                ).value
                if top_left_value is not None:
                    for row in range(merge_range.min_row, merge_range.max_row + 1):
                        for col in range(merge_range.min_col, merge_range.max_col + 1):
                            merged_values[(row, col)] = str(top_left_value)

            # Convert sheet to CSV text
            output = io.StringIO()
            writer = csv.writer(output)

            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                csv_row = []
                for col_idx, cell in enumerate(row, start=1):
                    # Check if this cell is part of a merged region
                    if (row_idx, col_idx) in merged_values:
                        csv_row.append(merged_values[(row_idx, col_idx)])
                    elif cell.value is not None:
                        csv_row.append(str(cell.value))
                    else:
                        csv_row.append("")
                # Skip completely empty rows
                if any(v.strip() for v in csv_row):
                    writer.writerow(csv_row)

            sheet_csv = output.getvalue().strip()
            if sheet_csv:
                sheets_text.append(f"--- Sheet: {sheet_name} ---\n{sheet_csv}")

        wb.close()
    except Exception as e:
        print(f"  Warning: Failed to extract text from {file_path}: {e}")
        return ""

    return "\n\n".join(sheets_text)


def _extract_text_from_xls(file_path: str) -> str:
    """Extract text from legacy .xls files using xlrd."""
    try:
        import xlrd
    except ImportError:
        print(f"  Warning: xlrd not installed. Cannot process {file_path}")
        return ""

    sheets_text: list[str] = []

    try:
        wb = xlrd.open_workbook(file_path)

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)

            output = io.StringIO()
            writer = csv.writer(output)

            for row_idx in range(ws.nrows):
                csv_row = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.value is not None and str(cell.value).strip():
                        csv_row.append(str(cell.value))
                    else:
                        csv_row.append("")
                if any(v.strip() for v in csv_row):
                    writer.writerow(csv_row)

            sheet_csv = output.getvalue().strip()
            if sheet_csv:
                sheets_text.append(f"--- Sheet: {sheet_name} ---\n{sheet_csv}")

    except Exception as e:
        print(f"  Warning: Failed to extract text from {file_path}: {e}")
        return ""

    return "\n\n".join(sheets_text)


def extract_text_from_csv(file_path: str) -> str:
    """Extract text from a CSV file, handling encoding issues."""
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, UnicodeError):
            continue
    # Last resort: read as binary and decode with replacement
    with open(file_path, "rb") as f:
        return f.read().decode("utf-8", errors="replace")


def extract_text(file_path: str) -> str:
    """Extract text from a supported file format."""
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return extract_text_from_pdf(file_path)
    elif ext in (".docx", ".doc"):
        return extract_text_from_docx(file_path)
    elif ext in (".xlsx", ".xls"):
        return extract_text_from_excel(file_path)
    elif ext == ".csv":
        return extract_text_from_csv(file_path)
    elif ext in (".txt", ".md"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    else:
        print(f"Skipping unsupported file type: {ext}")
        return ""


def get_embeddings(texts: list[str], batch_size: int = 32) -> list[list[float]]:
    """Generate embeddings using Azure OpenAI, with batching to avoid rate limits."""
    client = AzureOpenAI(
        azure_endpoint=settings.azure_openai_endpoint,
        api_key=settings.azure_openai_api_key,
        api_version=settings.azure_openai_api_version,
    )
    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i : i + batch_size]
        if i > 0:
            time.sleep(3)  # pace requests to stay under rate limit
        # Retry with exponential backoff — transient 401/429/timeout must not
        # abort a multi-thousand-file ingestion.
        for attempt in range(6):
            try:
                response = client.embeddings.create(
                    model=settings.azure_openai_embedding_deployment,
                    input=batch,
                )
                break
            except Exception as e:
                if attempt < 5:
                    wait = 2 ** attempt  # 1,2,4,8,16s
                    print(f"  Embed batch throttled/failed, retry in {wait}s... ({e})")
                    time.sleep(wait)
                else:
                    raise
        all_embeddings.extend([item.embedding for item in response.data])
        print(f"  Embedded {min(i + batch_size, len(texts))}/{len(texts)} chunks")
    return all_embeddings


def create_search_index():
    """Create the Azure AI Search index with vector fields."""
    index_client = SearchIndexClient(
        endpoint=settings.azure_search_endpoint,
        credential=AzureKeyCredential(settings.azure_search_api_key),
    )

    fields = [
        SimpleField(name="id", type=SearchFieldDataType.String, key=True),
        SearchableField(name="content", type=SearchFieldDataType.String),
        SimpleField(name="title", type=SearchFieldDataType.String, filterable=True),
        SimpleField(name="source", type=SearchFieldDataType.String, filterable=True),
        SearchField(
            name="content_vector",
            type=SearchFieldDataType.Collection(SearchFieldDataType.Single),
            searchable=True,
            vector_search_dimensions=1536,
            vector_search_profile_name="tibu-vector-profile",
        ),
    ]

    vector_search = VectorSearch(
        algorithms=[HnswAlgorithmConfiguration(name="tibu-hnsw")],
        profiles=[
            VectorSearchProfile(
                name="tibu-vector-profile",
                algorithm_configuration_name="tibu-hnsw",
            )
        ],
    )

    index = SearchIndex(
        name=settings.azure_search_index_name,
        fields=fields,
        vector_search=vector_search,
    )

    index_client.create_or_update_index(index)
    print(f"Index '{settings.azure_search_index_name}' created/updated.")


def ingest_documents(directory: str):
    """Process all documents in a directory and upload to Azure AI Search."""
    search_client = SearchClient(
        endpoint=settings.azure_search_endpoint,
        index_name=settings.azure_search_index_name,
        credential=AzureKeyCredential(settings.azure_search_api_key),
    )

    doc_id = 0
    failed_files: list[str] = []

    all_files = [
        p for p in Path(directory).rglob("*")
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    total_files = len(all_files)
    print(f"Found {total_files} file(s) to process in {directory}")

    def _upload_batch(batch: list) -> None:
        """Upload a single batch with retry + exponential backoff."""
        for attempt in range(5):
            try:
                search_client.upload_documents(batch)
                return
            except Exception as e:
                if attempt < 4:
                    wait = 2 ** attempt  # 1s, 2s, 4s, 8s
                    print(f"  Upload throttled, retrying in {wait}s... ({e})")
                    time.sleep(wait)
                else:
                    raise RuntimeError(f"Upload failed after 5 attempts — {e}") from e

    # Upload per-file so a mid-run crash keeps all prior progress, and one bad
    # file is skipped instead of aborting the whole (multi-thousand-file) run.
    # Re-runs are safe: chunk ids are deterministic, so they upsert.
    BATCH_SIZE = 100
    for file_idx, file_path in enumerate(all_files, 1):
        print(f"[{file_idx}/{total_files}] Processing: {file_path.name}")
        try:
            text = extract_text(str(file_path))
            if not text.strip():
                continue

            chunks = chunk_text(text)
            embeddings = get_embeddings(chunks)

            file_docs = []
            for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
                chunk_id = hashlib.md5(f"{file_path.name}:{i}".encode()).hexdigest()[:16]
                file_docs.append(
                    {
                        "id": f"doc-{chunk_id}",
                        "content": chunk,
                        "title": file_path.stem,
                        "source": file_path.name,
                        "content_vector": embedding,
                    }
                )

            for j in range(0, len(file_docs), BATCH_SIZE):
                _upload_batch(file_docs[j : j + BATCH_SIZE])
            doc_id += len(file_docs)
        except Exception as e:
            print(f"  SKIP {file_path.name}: {type(e).__name__}: {str(e)[:160]}")
            failed_files.append(file_path.name)
            continue

    print(f"Done! Ingested {doc_id} chunks from {directory}")
    if failed_files:
        print(f"Skipped {len(failed_files)} file(s) after retries: {failed_files[:20]}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Ingest documents into TIBU knowledge base")
    parser.add_argument("--path", default="./data/sample/", help="Directory containing documents")
    parser.add_argument("--create-index", action="store_true", help="Create/update the search index first")
    parser.add_argument("--index-only", action="store_true", help="Only create/update the index, skip ingestion")
    args = parser.parse_args()

    if args.create_index or args.index_only:
        create_search_index()

    if args.index_only:
        print("Index created. Skipping ingestion (--index-only).")
        exit(0)

    ingest_documents(args.path)
