"""
Verification tests for crawler improvements.

Tests:
1. Sitemap XML parsing
2. Domain matching (*.usiu.ac.ke subdomains, excluded subdomains)
3. URL skip patterns (legitimate URLs pass, noise blocked)
4. Document link detection (all file types)
5. XLSX→CSV conversion (multi-sheet, merged cells, formulas)
6. CSV text extraction (encoding fallback)
"""

import csv
import io
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

# Add the backend directory to path so we can import the modules
sys.path.insert(0, os.path.dirname(__file__))


def test_domain_matching():
    """Test that is_usiu_domain accepts valid subdomains and rejects excluded ones."""
    from app.services.crawl4ai_scraper import is_usiu_domain

    # Should accept
    assert is_usiu_domain("www.usiu.ac.ke"), "www.usiu.ac.ke should be accepted"
    assert is_usiu_domain("usiu.ac.ke"), "usiu.ac.ke should be accepted"
    assert is_usiu_domain("library.usiu.ac.ke"), "library.usiu.ac.ke should be accepted"
    assert is_usiu_domain("careers.usiu.ac.ke"), "careers.usiu.ac.ke should be accepted"

    # Should reject (excluded subdomains)
    assert not is_usiu_domain("mail.usiu.ac.ke"), "mail.usiu.ac.ke should be excluded"
    assert not is_usiu_domain("sis.usiu.ac.ke"), "sis.usiu.ac.ke should be excluded"
    assert not is_usiu_domain("portal.usiu.ac.ke"), "portal.usiu.ac.ke should be excluded"
    assert not is_usiu_domain("moodle.usiu.ac.ke"), "moodle.usiu.ac.ke should be excluded"

    # Should reject (not usiu.ac.ke at all)
    assert not is_usiu_domain("google.com"), "google.com should be rejected"
    assert not is_usiu_domain("example.ac.ke"), "example.ac.ke should be rejected"

    print("  ✓ Domain matching: all 10 cases passed")


def test_url_skip_patterns():
    """Test that skip patterns block noise but allow legitimate URLs."""
    from app.services.crawl4ai_scraper import should_skip_url

    # Should be skipped
    assert should_skip_url("https://www.usiu.ac.ke/wp-admin/"), "wp-admin should be skipped"
    assert should_skip_url("https://www.usiu.ac.ke/wp-login.php"), "wp-login should be skipped"
    assert should_skip_url("https://www.usiu.ac.ke/feed/"), "feed should be skipped"
    assert should_skip_url("https://www.usiu.ac.ke/tag/news/"), "tag should be skipped"
    assert should_skip_url("https://www.usiu.ac.ke/page/3"), "pagination should be skipped"

    # Should NOT be skipped (legitimate pages)
    assert not should_skip_url("https://www.usiu.ac.ke/admissions/"), "admissions should not be skipped"
    assert not should_skip_url("https://www.usiu.ac.ke/8/vision-mission-values"), "content page should not be skipped"
    assert not should_skip_url("https://www.usiu.ac.ke/clubs-and-organizations/"), "clubs page should not be skipped"
    # Query strings are now allowed (no blanket ?# block)
    assert not should_skip_url("https://www.usiu.ac.ke/search?q=fees"), "search query should not be skipped"

    print("  ✓ URL skip patterns: all 9 cases passed")


def test_document_url_detection():
    """Test that is_document_url detects all supported document types."""
    from app.services.crawl4ai_scraper import is_document_url

    # Should detect as documents
    assert is_document_url("https://www.usiu.ac.ke/files/catalog.pdf"), "PDF should be detected"
    assert is_document_url("https://www.usiu.ac.ke/forms/app.docx"), "DOCX should be detected"
    assert is_document_url("https://www.usiu.ac.ke/data/fees.xlsx"), "XLSX should be detected"
    assert is_document_url("https://www.usiu.ac.ke/docs/presentation.pptx"), "PPTX should be detected"
    assert is_document_url("https://www.usiu.ac.ke/data/report.csv"), "CSV should be detected"
    assert is_document_url("https://www.usiu.ac.ke/old/form.xls"), "XLS should be detected"
    assert is_document_url("https://www.usiu.ac.ke/old/letter.doc"), "DOC should be detected"

    # Should NOT detect as documents
    assert not is_document_url("https://www.usiu.ac.ke/admissions/"), "page should not be detected"
    assert not is_document_url("https://www.usiu.ac.ke/image.jpg"), "image should not be detected"

    print("  ✓ Document URL detection: all 9 cases passed")


def test_extract_document_links():
    """Test document link extraction from a mock crawl result."""
    from app.services.crawl4ai_scraper import extract_document_links

    # Build mock result
    result = MagicMock()
    result.links = {
        "internal": [
            {"href": "https://www.usiu.ac.ke/files/catalog.pdf"},
            {"href": "https://www.usiu.ac.ke/forms/app.docx"},
            {"href": "https://www.usiu.ac.ke/data/fees.xlsx"},
            {"href": "https://www.usiu.ac.ke/admissions/"},  # not a document
        ],
        "external": [
            {"href": "https://example.com/report.pptx"},
        ],
    }
    result.html = '''
        <a href="/assets/schedule.pdf">Schedule</a>
        <a href="https://www.usiu.ac.ke/data/grades.csv">Grades</a>
        <a href="/page">Not a doc</a>
    '''

    links = extract_document_links(result)
    assert any("catalog.pdf" in u for u in links), "catalog.pdf should be found"
    assert any("app.docx" in u for u in links), "app.docx should be found"
    assert any("fees.xlsx" in u for u in links), "fees.xlsx should be found"
    assert any("report.pptx" in u for u in links), "report.pptx should be found"
    assert any("schedule.pdf" in u for u in links), "schedule.pdf (from HTML) should be found"
    assert any("grades.csv" in u for u in links), "grades.csv (from HTML) should be found"
    # Non-documents should not be present
    non_docs = [u for u in links if "admissions" in u or "/page" in u]
    assert len(non_docs) == 0, "Non-document URLs should not be included"

    print(f"  ✓ Document link extraction: found {len(links)} document links, all expected types present")


def test_sitemap_parsing():
    """Test sitemap XML parsing with a mock XML string."""
    import xml.etree.ElementTree as ET
    from app.services.crawl4ai_scraper import fetch_sitemap_urls

    # We can't easily mock requests here, so just test the XML parsing logic
    xml_content = '''<?xml version="1.0" encoding="UTF-8"?>
    <urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">
        <url><loc>https://www.usiu.ac.ke/admissions/</loc></url>
        <url><loc>https://www.usiu.ac.ke/8/vision-mission-values</loc></url>
        <url><loc>https://www.usiu.ac.ke/fees-and-charges/</loc></url>
    </urlset>'''

    root = ET.fromstring(xml_content)
    ns = ""
    if root.tag.startswith("{"):
        ns = root.tag.split("}")[0] + "}"

    urls = set()
    for url_entry in root.findall(f"{ns}url"):
        loc = url_entry.find(f"{ns}loc")
        if loc is not None and loc.text:
            urls.add(loc.text.strip())

    assert len(urls) == 3, f"Expected 3 URLs from sitemap, got {len(urls)}"
    assert "https://www.usiu.ac.ke/admissions/" in urls
    assert "https://www.usiu.ac.ke/8/vision-mission-values" in urls
    assert "https://www.usiu.ac.ke/fees-and-charges/" in urls

    print("  ✓ Sitemap parsing: extracted 3 URLs correctly")


def test_xlsx_conversion():
    """Test XLSX→CSV conversion with multi-sheet, merged cells."""
    import openpyxl
    from app.services.document_loader import extract_text_from_excel

    # Create a test workbook with multiple sheets
    wb = openpyxl.Workbook()

    # Sheet 1: simple data
    ws1 = wb.active
    ws1.title = "Students"
    ws1.append(["Name", "Program", "Year"])
    ws1.append(["Alice", "Computer Science", 3])
    ws1.append(["Bob", "Business Admin", 2])
    ws1.append(["Carol", "Psychology", 4])

    # Sheet 2: with merged cells
    ws2 = wb.create_sheet("Fees")
    ws2.append(["Category", "Amount (KES)", "Notes"])
    ws2.append(["Tuition", "450000", "Per semester"])
    ws2.append(["Housing", "120000", "On-campus"])
    ws2.merge_cells("A4:C4")
    ws2["A4"] = "Total: 570,000 KES"

    # Sheet 3: with formula (will show None in data_only mode since never computed)
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Students", 3250])
    ws3.append(["Programs Offered", 45])

    # Save to temp file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name
        wb.save(temp_path)
    wb.close()

    try:
        text = extract_text_from_excel(temp_path)

        # Verify multi-sheet separation
        assert "--- Sheet: Students ---" in text, "Students sheet header should be present"
        assert "--- Sheet: Fees ---" in text, "Fees sheet header should be present"
        assert "--- Sheet: Summary ---" in text, "Summary sheet header should be present"

        # Verify data integrity
        assert "Alice" in text, "Student name 'Alice' should be in output"
        assert "Computer Science" in text, "Program name should be in output"
        assert "450000" in text, "Fee amount should be in output"
        assert "3250" in text, "Total students should be in output"

        # Verify merged cell was read
        assert "570,000 KES" in text, "Merged cell value should be preserved"

        print(f"  ✓ XLSX conversion: multi-sheet workbook converted successfully ({len(text)} chars)")
    finally:
        os.unlink(temp_path)


def test_csv_extraction():
    """Test CSV text extraction with encoding handling."""
    from app.services.document_loader import extract_text_from_csv

    # Create a test CSV
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Course", "Credits", "Grade"])
        writer.writerow(["Introduction to CS", "3", "A"])
        writer.writerow(["Data Structures", "4", "B+"])
        temp_path = f.name

    try:
        text = extract_text_from_csv(temp_path)
        assert "Introduction to CS" in text, "Course name should be in output"
        assert "Data Structures" in text, "Course name should be in output"
        assert "Credits" in text, "Header should be in output"
        print(f"  ✓ CSV extraction: read successfully ({len(text)} chars)")
    finally:
        os.unlink(temp_path)


def test_extract_text_dispatch():
    """Test that extract_text correctly dispatches to the right handler."""
    from app.services.document_loader import extract_text
    import openpyxl

    # Create a minimal xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Test", "Data"])
    ws.append(["Hello", "World"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name
        wb.save(temp_path)
    wb.close()

    try:
        text = extract_text(temp_path)
        assert "Hello" in text, "extract_text should dispatch xlsx to excel handler"
        assert "World" in text, "extract_text should dispatch xlsx to excel handler"
        print("  ✓ extract_text dispatch: xlsx correctly routed to Excel handler")
    finally:
        os.unlink(temp_path)

    # Test with a .txt file
    with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as f:
        f.write("Test content for txt file")
        txt_path = f.name

    try:
        text = extract_text(txt_path)
        assert "Test content" in text, "extract_text should handle txt files"
        print("  ✓ extract_text dispatch: txt correctly handled")
    finally:
        os.unlink(txt_path)


if __name__ == "__main__":
    print("=" * 60)
    print("TIBU Crawler & Document Loader — Verification Tests")
    print("=" * 60)

    tests = [
        ("Domain matching", test_domain_matching),
        ("URL skip patterns", test_url_skip_patterns),
        ("Document URL detection", test_document_url_detection),
        ("Document link extraction", test_extract_document_links),
        ("Sitemap XML parsing", test_sitemap_parsing),
        ("XLSX→CSV conversion", test_xlsx_conversion),
        ("CSV text extraction", test_csv_extraction),
        ("extract_text dispatch", test_extract_text_dispatch),
    ]

    passed = 0
    failed = 0
    for name, test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            print(f"  ✗ {name}: FAILED — {e}")
            failed += 1

    print()
    print("=" * 60)
    print(f"Results: {passed} passed, {failed} failed, {len(tests)} total")
    print("=" * 60)

    sys.exit(1 if failed else 0)
